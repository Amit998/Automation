import openpyxl



# Fetch The Sheet

# fileName="test1.xlsx"
# workbook=openpyxl.load_workbook(fileName)
# sheet=workbook.active
# sheet.title="Changed"


# print(sheet['A1'].value)
# print(sheet['A2'].value)
# print(sheet['A3'].value)
# print(sheet['B1'].value)
# print(sheet['B2'].value)
# print(sheet['B3'].value)


# sheet['A1']="Amit"
# sheet['C1']="B2+B3"



# workbook.save(fileName)


# Create New Sheet


# fileName="test1.xlsx"
# workbook=openpyxl.load_workbook(fileName)
# sheet=workbook.active
# sheet.title="Changed"


# workbook.create_sheet("Second")
# sheet=workbook['Second']
# sheet['A1']="What is up"
# workbook.save(fileName)



# workbook=openpyxl.Workbook()
# sheet=workbook.active

# sheet['A1']=10
# sheet['B1']=10
# sheet['C1']=10
# sheet['D1']=10
# sheet['E1']=10

# sheet.insert_rows(1,amount=5)
# sheet.insert_cols(2,amount=5)


# sheet.merge_cells("A1:C3")
# sheet['A1']="Hello World"

# sheet.move_range("A1:C3",rows=6,cols=4)

# sheet.move_range("A1:C3",rows=-3,cols=-1)

# workbook.save("newBook.xlsx")



filename="file_example_XLSX_10.xlsx"
import datetime as dt
import pandas as pd

workbook=openpyxl.load_workbook(filename)

sheet=workbook.active

for row in range(1,11):
    date=sheet[f'G{row}'].value
   
    today=dt.datetime.now()

    

    # print(date)
    dateYear=(date[-4:])
    dateDay=(date[:2])
    dateMonth=(date[3:5])



    
    age=today.year - dateYear - (( today.month,today.day ) < (date.month,today.day) )
    sheet[f'I{row}']=age

workbook.save(filename)